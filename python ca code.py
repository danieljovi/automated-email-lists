
#Name :Feguem COrine
#Matricule: 20212150


import openpyxl as xl

wb = xl.load_workbook('employeedata.xlsx')
sheet = wb['sheet1'] #sheet1 is the first file i created with the email lists as as helpinghands.cm
cell = sheet.cell(2,3) #reads the entry in row1,column2
print(cell.value)

row = sheet.max_row
column = sheet.max_column #this defines the rows and column

#creating the new row for updated list#
for row in range (2, sheet.max_row +1): #begins from 2 because we are working with 2nd row elements
     cell = sheet.cell(row, 1) #reads all element of column 1
     new_email = (cell.value + '@helpinghands.org') #cuncatenates value in cell with @helpinghands.org
                                                    #creates and stored new email
new_email_cell = sheet.cell(row,4) #stores the new email im column E
new_email_cell.value = new_email

#creating a csv file that sores the email
first_domain = 'helpinghands.cm'
second_domain = 'helpinghands.org'
for i in range(2,sheet.max_row + 1);
cell = sheet.cell(i,3)
  if first_domain in cell.value:
    updated_email=(cell.value).replace(first_domain,second_domain)

    sheet.cell(i,4).value = upgrated_email
 wb.save( 'updated_emails.csv')   